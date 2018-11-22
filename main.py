import util                                                         #For Calculate Differentiation
import random                                                       #For Parameter Initialization
from openpyxl import load_workbook                                  #For Loading .xlsx file

EPOCHS = 100                                                        #Linear Regression Epochs
LEARNING_RATE = 1e-3                                                #Linear Regression Learning Rate

Util = util.Util(LEARNING_RATE)                                     #Initialize Util
random.seed(8)
X, Y = [], []                                                       #Create X and Y Data

filepath = 'data.xlsx'                                              #Data's Location
xl = load_workbook(filepath)                                        #Load .xlsx file
sheet = xl.active                                                   #Load Lastly Sheet
max_row = sheet.max_row                                             #Get Data's Row
for i in range(1, max_row + 1):                                     #Append Data in Each Array
    X.append(sheet.cell(row=i, column=1).value)
    Y.append(sheet.cell(row=i, column=2).value)

Util.input_data(X, Y)                                               #Input Data to Util Class

w = random.random()                                                 #Create Weight
b = random.random()                                                 #Crate Bias

for i in range(EPOCHS):
    cost = 0                                                        #Calculate Cost
    for j in range(max_row):
        hy = w * X[j] + b                                           #Linear Equation
        cost = cost + (Y[j] - hy) ** 2                              #Cost Formula
    cost = cost / max_row

    w = w + Util.Differentiation(w, b, max_row) * LEARNING_RATE     #Calculate Differentiation in Util Class and multiply with LEARNING_RATE, Cause the Differentiation has big value.
    b = b + Util.Differentiation(w, b, max_row) * LEARNING_RATE     #Calculate Differentiation in Util Class and multiply with LEARNING_RATE, Cause the Differentiation has big value.

    if i % 10 == 0:
        print('EPOCH %d - loss %f' % (i, cost))

print('w is %f' % w)
print('b is %f' % b)

x = input("Enter the number what you want! ")
print("Result is %f" % (w * int(x) + b))